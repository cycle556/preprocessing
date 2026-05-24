# -*- coding: utf-8 -*-
"""Phase 3: 多层分块器 (chunker v2)

分层策略：
  Layer 0: doc_summary — 每文档 1 条，含标题+章节列表
  Layer 1: section      — 每 H2 主题 1 条，含章节摘要+子块列表
  Layer 2: leaf         — 内容块，≤ MAX_CHARS
  Layer 3: table_row    — 大表每行 1 条，带完整表头

"""

from __future__ import annotations

import json
import os
import re
from collections import Counter
from dataclasses import dataclass, asdict, field
from typing import List, Tuple, Optional

from lib.utils import count_tokens, sanitize_for_id

# --- 默认参数 ---
DEFAULT_MAX_CHARS = 1800
DEFAULT_MIN_CHARS = 80
DEFAULT_OVERLAP = 80

RECURSIVE_SEPARATORS = [
    "\n\n", "\n",
    "。", "！", "？",
    ". ", "! ", "? ",
    "；", "; ",
    "，", ", ",
    " ", "",
]


# ============================================================================
# 数据结构
# ============================================================================

@dataclass
class Chunk:
    """兼容现有 JSONL schema，新增分层检索字段。"""

    # --- 现有字段 ---
    chunk_id: str
    source: str
    breadcrumb: List[str] = field(default_factory=list)
    section_title: str = ""
    heading_level: int = 0
    content: str = ""
    content_with_context: str = ""
    char_count: int = 0
    est_token_count: int = 0
    has_table: bool = False
    table_caption: Optional[str] = None
    is_split_part: bool = False
    split_part_index: Optional[int] = None
    split_part_total: Optional[int] = None
    page_start: Optional[int] = None
    page_end: Optional[int] = None

    # --- 新增分层字段 ---
    chunk_type: str = "leaf"       # doc_summary | section | leaf | table_row
    parent_id: Optional[str] = None
    children_ids: List[str] = field(default_factory=list)
    sibling_ids: List[str] = field(default_factory=list)
    content_type: str = "prose"    # prose | table | mixed
    company: str = ""              # 所属保险公司


# ============================================================================
# Markdown 解析（复用现有逻辑）
# ============================================================================

HEADING_RE = re.compile(r'^(#{1,6})\s+(.+?)\s*$')
CODE_FENCE_RE = re.compile(r'^\s*```')


def parse_sections(md_text: str) -> List[Tuple[int, str, str]]:
    """将 Markdown 拆分为 (标题级别, 标题文本, 正文) 三元组。"""
    sections: List[Tuple[int, str, str]] = []
    lines = md_text.splitlines()
    cur_level, cur_heading, cur_body = 0, "", []
    in_code = False

    for line in lines:
        if CODE_FENCE_RE.match(line):
            in_code = not in_code
            cur_body.append(line)
            continue
        if not in_code:
            m = HEADING_RE.match(line)
            if m:
                body_text = "\n".join(cur_body).strip("\n")
                if cur_level != 0 or body_text:
                    sections.append((cur_level, cur_heading, body_text))
                cur_level = len(m.group(1))
                cur_heading = m.group(2).strip()
                cur_body = []
                continue
        cur_body.append(line)

    body_text = "\n".join(cur_body).strip("\n")
    if cur_level != 0 or body_text:
        sections.append((cur_level, cur_heading, body_text))
    return sections


def build_breadcrumbs(sections: List[Tuple[int, str, str]]) -> List[Tuple[List[str], int, str, str]]:
    """为每个 section 附加面包屑导航。"""
    out = []
    stack: List[Tuple[int, str]] = []
    for level, heading, body in sections:
        if level == 0:
            out.append(([], 0, "", body))
            continue
        while stack and stack[-1][0] >= level:
            stack.pop()
        stack.append((level, heading))
        breadcrumb = [h for _, h in stack]
        out.append((breadcrumb, level, heading, body))
    return out


def detect_primary_level(walked) -> int:
    """自动检测主要章节边界层级。"""
    levels = [lvl for _, lvl, _, _ in walked if lvl > 0]
    if not levels:
        return 2
    counter = Counter(levels)
    for lvl in sorted(counter.keys()):
        if counter[lvl] >= 2:
            return lvl
    return min(counter.keys())


def _extract_document_title(sections) -> Optional[str]:
    """提取文档标题：第一个 H1。"""
    for level, heading, _ in sections:
        if level == 1 and heading.strip():
            return heading.strip()
    return None


# ============================================================================
# 表格处理
# ============================================================================

def _is_table_separator(line: str) -> bool:
    s = line.strip().strip('|').strip()
    if not s:
        return False
    parts = [p.strip() for p in s.split('|')]
    return len(parts) >= 2 and all(re.fullmatch(r':?-{2,}:?', p) for p in parts if p)


def _contains_table(text: str) -> bool:
    lines = text.split('\n')
    for i in range(len(lines) - 1):
        if '|' in lines[i] and _is_table_separator(lines[i + 1]):
            return True
    return False


def _extract_table_header(table_text: str) -> Tuple[Optional[str], Optional[str]]:
    """提取表格的表头行和分隔行。"""
    lines = table_text.split('\n')
    for i in range(len(lines) - 1):
        if '|' in lines[i] and _is_table_separator(lines[i + 1]):
            return lines[i], lines[i + 1]
    return None, None


def split_table_by_rows(table_text: str, max_chars: int) -> list[str]:
    """按数据行拆分表格；每组尽量塞满 max_chars，每组拷贝表头+分隔行。"""
    lines = [ln for ln in table_text.split('\n') if ln.strip()]
    if len(lines) < 3:
        return [table_text]
    header, sep = lines[0], lines[1]
    data_rows = lines[2:]
    head_size = len(header) + len(sep) + 2
    out, cur, cur_size = [], [header, sep], head_size
    for row in data_rows:
        row_size = len(row) + 1
        if cur_size + row_size > max_chars and len(cur) > 2:
            out.append('\n'.join(cur))
            cur, cur_size = [header, sep], head_size
        cur.append(row)
        cur_size += row_size
    if len(cur) > 2:
        out.append('\n'.join(cur))
    return out or [table_text]


# ============================================================================
# 递归字符切分（prose 回退）
# ============================================================================

def recursive_char_split(text: str, max_chars: int, overlap: int,
                         seps: list = RECURSIVE_SEPARATORS) -> list[str]:
    """按由粗到细的顺序使用分隔符，超限时递归降级。"""
    if len(text) <= max_chars:
        return [text]
    for sep in seps:
        if sep == "":
            return _hard_char_split(text, max_chars, overlap)
        if sep not in text:
            continue
        pieces = _split_with_sep(text, sep)
        refined = []
        for p in pieces:
            refined.extend(
                recursive_char_split(p, max_chars, overlap,
                                     seps[seps.index(sep) + 1:])
                if len(p) > max_chars else [p]
            )
        return _merge_with_overlap(refined, max_chars, overlap)
    return _hard_char_split(text, max_chars, overlap)


def _split_with_sep(text: str, sep: str) -> list[str]:
    if not sep:
        return [text]
    parts = text.split(sep)
    return [p + (sep if i < len(parts) - 1 else "")
            for i, p in enumerate(parts) if p != ""]


def _merge_with_overlap(pieces: list[str], max_chars: int, overlap: int) -> list[str]:
    out, cur = [], ""
    for p in pieces:
        if not cur:
            cur = p
        elif len(cur) + len(p) <= max_chars:
            cur += p
        else:
            out.append(cur)
            tail = cur[-overlap:] if overlap and len(cur) > overlap else ""
            cur = tail + p
    if cur:
        out.append(cur)
    return out


def _hard_char_split(text: str, max_chars: int, overlap: int) -> list[str]:
    out, i = [], 0
    while i < len(text):
        out.append(text[i:i + max_chars])
        i += max(1, max_chars - overlap)
    return out


# ============================================================================
# 语义切分
# ============================================================================

def _lazy_load_embedding():
    """懒加载 embedding 模型 —— 在需要语义切分时才加载。"""
    import os
    try:
        from sentence_transformers import SentenceTransformer
        model_path = os.environ.get(
            "EMBEDDING_MODEL_PATH",
            r"C:\Users\CYL\Desktop\finnal\models\BAAI\bge-m3"
        )
        return SentenceTransformer(model_path, local_files_only=True)
    except Exception:
        return None


def semantic_split(paragraphs: list[str], embedding_model=None,
                   threshold: float = 0.6) -> list[str]:
    """在语义断裂处切分段落序列。

    Args:
        paragraphs: 段落文本列表
        embedding_model: SentenceTransformer 实例（None 时回退到 recursive）
        threshold: 相似度低于(局部均值 × threshold)时在此切分
    """
    if embedding_model is None or len(paragraphs) <= 1:
        text = '\n\n'.join(paragraphs)
        return [text]

    try:
        import numpy as np
        vecs = embedding_model.encode(paragraphs)
        # 余弦相似度
        norms = np.linalg.norm(vecs, axis=1, keepdims=True)
        norms[norms == 0] = 1
        vecs_norm = vecs / norms
        sims = [(vecs_norm[i] @ vecs_norm[i + 1]) for i in range(len(vecs_norm) - 1)]

        breakpoints = [0]
        window = 3
        for i, s in enumerate(sims):
            start = max(0, i - window)
            end = min(len(sims), i + window + 1)
            local_mean = float(np.mean(sims[start:end]))
            if local_mean > 0 and s < local_mean * threshold:
                breakpoints.append(i + 1)
        breakpoints.append(len(paragraphs))

        chunks = []
        for a, b in zip(breakpoints, breakpoints[1:]):
            if a < b:
                chunks.append('\n\n'.join(paragraphs[a:b]))
        return chunks if chunks else ['\n\n'.join(paragraphs)]
    except Exception:
        text = '\n\n'.join(paragraphs)
        return [text]


# ============================================================================
# 正文 → 内容块分解
# ============================================================================

def split_text_into_blocks(text: str) -> list[tuple[str, str]]:
    """将章节正文拆分为 [(类型, 内容)]，类型为 'table' 或 'prose'。"""
    lines = text.split('\n')
    blocks, i, n = [], 0, len(lines)
    while i < n:
        if '|' in lines[i] and i + 1 < n and _is_table_separator(lines[i + 1]):
            j = i + 2
            while j < n and '|' in lines[j] and lines[j].strip():
                j += 1
            blocks.append(('table', '\n'.join(lines[i:j])))
            i = j
        else:
            j = i
            while j < n:
                if (j + 1 < n and '|' in lines[j]
                        and _is_table_separator(lines[j + 1])):
                    break
                j += 1
            prose = '\n'.join(lines[i:j]).strip('\n')
            if prose:
                blocks.append(('prose', prose))
            i = j
    return blocks


# ============================================================================
# 核心：多层分块主逻辑
# ============================================================================

def chunk_markdown_v2(
    md_text: str,
    source: str,
    max_chars: int = DEFAULT_MAX_CHARS,
    overlap: int = DEFAULT_OVERLAP,
    id_prefix: str = "chunk",
    min_chars: int = DEFAULT_MIN_CHARS,
    embedding_model=None,
    company: str = "",
) -> list[Chunk]:
    """多层分块：doc_summary → section → leaf/table_row。

    Args:
        md_text: 清洗+层级化后的 Markdown
        source: 源文件名
        max_chars: 叶子块最大字符数
        overlap: prose 子块重叠字符数
        id_prefix: chunk_id 前缀
        min_chars: 过短块告警阈值
        embedding_model: 可选，SentenceTransformer 实例，用于语义切分
    """
    sections = parse_sections(md_text)
    walked = build_breadcrumbs(sections)
    doc_title = _extract_document_title(sections)
    primary_level = detect_primary_level(walked)

    all_chunks: list[Chunk] = []
    counter = [0]  # mutable counter

    def _next_id() -> str:
        counter[0] += 1
        return f"{id_prefix}_{counter[0]:04d}"

    # --- Layer 0: doc_summary ---
    doc_id = _next_id()
    # 提取文档摘要：标题 + 前 300 字符
    first_body = sections[0][2] if sections else ""
    doc_summary_text = (doc_title or source) + "\n\n" + first_body[:300]
    doc_bc = [doc_title] if doc_title else []

    # 先遍历一次收集所有 H2 标题名
    h2_headings = []
    for breadcrumb, level, heading, body in walked:
        if level == primary_level and heading.strip():
            h2_headings.append(heading.strip())

    doc_content = (doc_title or source) + "\n\n"
    if h2_headings:
        doc_content += "本文件包含以下章节:\n" + "\n".join(f"- {h}" for h in h2_headings)

    doc_chunk = Chunk(
        chunk_id=doc_id, source=source,
        breadcrumb=doc_bc, section_title=doc_title or source,
        heading_level=1,
        content=doc_content,
        content_with_context=f"# {doc_title or source}\n\n{doc_content}",
        char_count=len(doc_content),
        est_token_count=count_tokens(doc_content),
        chunk_type="doc_summary",
        content_type="prose",
        company=company,
    )

    # --- 按 primary_level 合并小章节 ---
    merged = _merge_small_sections(walked, max_chars, primary_level)

    # 用于收集 section → leaf 的映射
    section_children: dict[str, list[str]] = {doc_id: []}

    for breadcrumb, level, heading, body in merged:
        if not body.strip():
            continue

        # 注入文档标题
        if doc_title and (not breadcrumb or breadcrumb[0] != doc_title):
            full_bc = [doc_title] + breadcrumb
        else:
            full_bc = list(breadcrumb)

        sec_title = heading.strip()

        # 判断是否为"表格型章节"
        is_table_section = _contains_table(body)

        # --- 检查 body 是否短到可以合入 section ---
        body_fits_in_one = len(body) <= max_chars
        if body_fits_in_one:
            # body 短：section 自己就是 leaf，不再创建子 leaf
            sec_id = _next_id()
            section_children[doc_id].append(sec_id)
            cw = f"# {' > '.join(full_bc)}\n\n{body}" if body.strip() else ""
            sec_chunk = Chunk(
                chunk_id=sec_id, source=source,
                breadcrumb=full_bc, section_title=sec_title,
                heading_level=level,
                content=body,
                content_with_context=cw,
                char_count=len(body),
                est_token_count=count_tokens(cw),
                has_table=is_table_section,
                table_caption=sec_title if is_table_section else None,
                chunk_type="section",
                parent_id=doc_id,
                content_type="table" if is_table_section else "prose",
                company=company,
            )
            all_chunks.append(sec_chunk)
            section_children[sec_id] = []  # 无子 leaf
            continue

        # --- body 长：创建 section 摘要 + leaf 子块 ---
        sec_summary = _make_section_summary(body, sec_title, is_table_section)
        sec_id = _next_id()
        section_children[doc_id].append(sec_id)

        sec_chunk = Chunk(
            chunk_id=sec_id, source=source,
            breadcrumb=full_bc, section_title=sec_title,
            heading_level=level,
            content=sec_summary,
            content_with_context=(
                f"# {' > '.join(full_bc)}\n\n{sec_summary}"
            ),
            char_count=len(sec_summary),
            est_token_count=count_tokens(sec_summary),
            has_table=is_table_section,
            chunk_type="section",
            parent_id=doc_id,
            content_type="table" if is_table_section else "prose",
            company=company,
        )
        all_chunks.append(sec_chunk)
        section_children[sec_id] = []

        # --- Layer 2 & 3: leaf / table_row ---
        blocks = split_text_into_blocks(body)
        if not blocks:
            sec_chunk.chunk_type = "leaf"
            sec_chunk.parent_id = doc_id
            continue

        leaf_ids: list[str] = []
        for kind, blk_text in blocks:
            if kind == 'table':
                leaf_ids += _chunk_table(
                    blk_text, max_chars, sec_id, sec_title, source,
                    full_bc, level, _next_id, all_chunks, company,
                )
            else:
                leaf_ids += _chunk_prose(
                    blk_text, max_chars, overlap, sec_id, sec_title, source,
                    full_bc, level, _next_id, all_chunks, embedding_model, company,
                )
        section_children[sec_id] = leaf_ids

    # --- 回填 children_ids ---
    for chunk in all_chunks:
        if chunk.chunk_id in section_children:
            chunk.children_ids = section_children[chunk.chunk_id]
    # 回填 doc_summary 的 children
    if doc_id in section_children:
        doc_chunk.children_ids = section_children[doc_id]

    # 将所有 chunk 放入列表（doc first, then sections+leaves）
    result = [doc_chunk] + all_chunks
    return result


def _make_section_summary(body: str, section_title: str, has_table: bool) -> str:
    """生成章节摘要，在干净边界处截断。

    规则：
    1. 优先在段落边界（空行）截断，不在句子或表格行中间切断
    2. 摘要控制在 200-350 字符之间
    3. 表格型章节在开头标注
    """
    max_preview = 300
    if len(body) <= max_preview:
        return body

    # 在 max_preview 附近找最近的空行边界
    lines = body.split('\n')
    best = min(max_preview, len(body))
    for i, line in enumerate(lines):
        pos = sum(len(l) + 1 for l in lines[:i])
        if pos > max_preview:
            # 回退到前一个空行或标题行
            for j in range(i - 1, max(0, i - 8), -1):
                prev = lines[j].strip()
                if prev == '' or prev.startswith('#'):
                    best = sum(len(l) + 1 for l in lines[:j])
                    break
            break

    preview = body[:best].strip()
    if has_table:
        preview = f"本章节包含表格数据: {section_title}\n\n{preview}"
    return preview


def _chunk_table(
    table_text: str, max_chars: int, parent_id: str,
    section_title: str, source: str, breadcrumb: list,
    level: int, next_id, all_chunks: list,
    company: str = "",
) -> list[str]:
    """处理表格：小表 → 1 leaf，大表 → 摘要 leaf + N row chunks。"""
    ctx_header = " > ".join(breadcrumb) if breadcrumb else "(preamble)"
    table_len = len(table_text)

    if table_len <= max_chars:
        # 小表 → 整表作为一个 leaf
        leaf_id = next_id()
        cw = f"# {ctx_header}\n\n{table_text}"
        all_chunks.append(Chunk(
            chunk_id=leaf_id, source=source,
            breadcrumb=list(breadcrumb), section_title=section_title,
            heading_level=level,
            content=table_text, content_with_context=cw,
            char_count=table_len,
            est_token_count=count_tokens(cw),
            has_table=True,
            table_caption=section_title,
            chunk_type="leaf", parent_id=parent_id,
            content_type="table",
            company=company,
        ))
        return [leaf_id]

    # 大表 → 拆分为子表 + 行级 chunks
    sub_tables = split_table_by_rows(table_text, max_chars)
    header, sep = _extract_table_header(table_text)

    leaf_ids = []
    sibling_ids_so_far = []
    total = len(sub_tables)

    for idx, sub in enumerate(sub_tables, start=1):
        leaf_id = next_id()
        leaf_ids.append(leaf_id)
        cw = f"# {ctx_header}\n\n{sub}"
        is_split = total > 1
        caption = f"{section_title}（子表 {idx}/{total}）" if is_split else section_title

        chunk = Chunk(
            chunk_id=leaf_id, source=source,
            breadcrumb=list(breadcrumb), section_title=section_title,
            heading_level=level,
            content=sub, content_with_context=cw,
            char_count=len(sub),
            est_token_count=count_tokens(cw),
            has_table=True, table_caption=caption,
            is_split_part=is_split,
            split_part_index=idx if is_split else None,
            split_part_total=total if is_split else None,
            chunk_type="leaf", parent_id=parent_id,
            content_type="table",
            company=company,
        )

        # 行级切分：同时为每行生成 table_row chunk
        if header and is_split:
            sub_lines = [l for l in sub.split('\n') if l.strip() and '---' not in l]
            data_lines = sub_lines[1:] if sub_lines else []  # skip header
            row_siblings = []
            for ri, row_line in enumerate(data_lines, start=1):
                if not row_line.strip():
                    continue
                row_id = next_id()
                row_siblings.append(row_id)
                row_content = f"{header}\n{sep}\n{row_line}"
                all_chunks.append(Chunk(
                    chunk_id=row_id, source=source,
                    breadcrumb=list(breadcrumb), section_title=section_title,
                    heading_level=level,
                    content=row_content,
                    content_with_context=f"# {ctx_header}\n\n{row_content}",
                    char_count=len(row_content),
                    est_token_count=count_tokens(row_content),
                    has_table=True, table_caption=f"{section_title} 行 {ri}",
                    chunk_type="table_row", parent_id=leaf_id,
                    content_type="table",
                    company=company,
                ))
            sibling_ids_so_far.append(row_siblings)

        all_chunks.append(chunk)

    # 回填 sibling_ids（同一表格的所有 leaf 子块 + 同一 leaf 下的行）
    table_leaf_indices = [i for i, c in enumerate(all_chunks)
                          if c.chunk_id in leaf_ids]
    for i in table_leaf_indices:
        all_chunks[i].sibling_ids = [lid for lid in leaf_ids if lid != all_chunks[i].chunk_id]

    return leaf_ids


def _chunk_prose(
    text: str, max_chars: int, overlap: int,
    parent_id: str, section_title: str, source: str,
    breadcrumb: list, level: int, next_id, all_chunks: list,
    embedding_model=None, company: str = "",
) -> list[str]:
    """处理纯文本：优先语义切分，回退递归切分。"""
    ctx_header = " > ".join(breadcrumb) if breadcrumb else "(preamble)"
    text_len = len(text)

    if text_len <= max_chars:
        leaf_id = next_id()
        cw = f"# {ctx_header}\n\n{text}"
        all_chunks.append(Chunk(
            chunk_id=leaf_id, source=source,
            breadcrumb=list(breadcrumb), section_title=section_title,
            heading_level=level,
            content=text, content_with_context=cw,
            char_count=text_len,
            est_token_count=count_tokens(cw),
            chunk_type="leaf", parent_id=parent_id,
            content_type="prose",
            company=company,
        ))
        return [leaf_id]

    # 尝试语义切分
    paragraphs = [p.strip() for p in text.split('\n\n') if p.strip()]
    if len(paragraphs) > 1:
        model = embedding_model or _lazy_load_embedding()
        if model is not None:
            semantic_chunks = semantic_split(paragraphs, model)
        else:
            semantic_chunks = [text]
    else:
        semantic_chunks = [text]

    leaf_ids = []
    for sc in semantic_chunks:
        if len(sc) <= max_chars:
            sub_texts = [sc]
        else:
            sub_texts = recursive_char_split(sc, max_chars, overlap)

        for sub in sub_texts:
            leaf_id = next_id()
            leaf_ids.append(leaf_id)
            cw = f"# {ctx_header}\n\n{sub}"
            all_chunks.append(Chunk(
                chunk_id=leaf_id, source=source,
                breadcrumb=list(breadcrumb), section_title=section_title,
                heading_level=level,
                content=sub, content_with_context=cw,
                char_count=len(sub),
                est_token_count=count_tokens(cw),
                chunk_type="leaf", parent_id=parent_id,
                content_type="prose",
                company=company,
            ))
    return leaf_ids


def _merge_small_sections(walked, max_chars: int, primary_level: int):
    """合并小主章节，避免碎片化。"""
    if not walked:
        return []
    primary_indices = [i for i, (_, lvl, _, _) in enumerate(walked)
                       if lvl == primary_level]
    out = []
    first_primary = primary_indices[0] if primary_indices else len(walked)
    for item in walked[:first_primary]:
        if item[1] == 0 and not item[3].strip():
            continue
        out.append(item)

    for k, start in enumerate(primary_indices):
        end = primary_indices[k + 1] if k + 1 < len(primary_indices) else len(walked)
        group = walked[start:end]
        p_bc, p_lvl, p_h, p_body = group[0]

        merged_body_parts = []
        if p_body.strip():
            merged_body_parts.append(p_body)
        for sub in group[1:]:
            _, sub_lvl, sub_h, sub_b = sub
            hashes = "#" * sub_lvl
            merged_body_parts.append(f"{hashes} {sub_h}\n{sub_b}".rstrip())
        merged_body = "\n\n".join(merged_body_parts).strip()

        if len(merged_body) <= max_chars or len(group) == 1:
            out.append((p_bc, p_lvl, p_h, merged_body))
        else:
            if p_body.strip():
                out.append((p_bc, p_lvl, p_h, p_body))
            for sub in group[1:]:
                out.append(sub)
    return out


# ============================================================================
# 输出
# ============================================================================

def _chunk_to_dict(chunk: Chunk) -> dict:
    """转换为 JSON 字典，供 embed_all.py / upload_polardb.py 等下游使用。"""
    return {
        "chunk_id": chunk.chunk_id,
        "source": chunk.source,
        "breadcrumb": chunk.breadcrumb,
        "section_title": chunk.section_title,
        "heading_level": chunk.heading_level,
        "content": chunk.content,
        "content_with_context": chunk.content_with_context,
        "char_count": chunk.char_count,
        "est_token_count": chunk.est_token_count,
        "has_table": chunk.has_table,
        "table_caption": chunk.table_caption or "",
        "is_split_part": chunk.is_split_part,
        "split_part_index": chunk.split_part_index,
        "split_part_total": chunk.split_part_total,
        "page_start": chunk.page_start,
        "page_end": chunk.page_end,
        # 新增字段
        "chunk_type": chunk.chunk_type,
        "parent_id": chunk.parent_id or "",
        "children_ids": chunk.children_ids,
        "sibling_ids": chunk.sibling_ids,
        "content_type": chunk.content_type,
        "company": chunk.company,
    }


def write_jsonl(chunks: list[Chunk], path: str) -> None:
    """输出 JSONL。"""
    with open(path, "w", encoding="utf-8") as f:
        for c in chunks:
            f.write(json.dumps(_chunk_to_dict(c), ensure_ascii=False) + "\n")


def write_xlsx(chunks: list[Chunk], path: str) -> None:
    """输出 XLSX 用于人工审阅切分质量。"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("  [warn] openpyxl 未安装，跳过 XLSX 输出（pip install openpyxl）")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "chunks"

    headers = [
        "chunk_id", "chunk_type", "source", "company",
        "breadcrumb", "section_title", "heading_level",
        "char_count", "est_token_count", "has_table", "table_caption",
        "is_split_part", "split_part_index", "split_part_total",
        "parent_id", "children_ids", "sibling_ids",
        "content_type", "content", "content_with_context",
    ]
    ws.append(headers)

    for c in chunks:
        ws.append([
            c.chunk_id, c.chunk_type, c.source, c.company,
            " > ".join(c.breadcrumb), c.section_title, c.heading_level,
            c.char_count, c.est_token_count, c.has_table,
            c.table_caption or "",
            c.is_split_part,
            c.split_part_index if c.split_part_index is not None else "",
            c.split_part_total if c.split_part_total is not None else "",
            c.parent_id or "",
            ", ".join(c.children_ids) if c.children_ids else "",
            ", ".join(c.sibling_ids) if c.sibling_ids else "",
            c.content_type, c.content, c.content_with_context,
        ])

    # 格式
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", start_color="305496")
    header_align = Alignment(horizontal="center", vertical="center")
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    widths = {
        "chunk_id": 18, "chunk_type": 12, "source": 24, "company": 10,
        "breadcrumb": 40, "section_title": 24, "heading_level": 8,
        "char_count": 9, "est_token_count": 10, "has_table": 8,
        "is_split_part": 9, "split_part_index": 9, "split_part_total": 9,
        "parent_id": 18, "content": 50, "content_with_context": 60,
    }
    for i, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(h, 15)

    wrap_cols = {headers.index(h) + 1 for h in
                 ("breadcrumb", "content", "content_with_context",
                  "children_ids", "sibling_ids")}
    body_font = Font(name="Arial", size=10)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = body_font
            cell.alignment = (
                Alignment(wrap_text=True, vertical="top")
                if cell.column in wrap_cols else
                Alignment(vertical="top")
            )

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(path)


def write_md(chunks, path: str, min_chars: int = 80) -> None:
    """输出 Markdown 审阅文件。"""
    if not chunks:
        with open(path, "w", encoding="utf-8") as f:
            f.write("# Chunks: (empty)\n\nNo chunks produced.\n")
        return

    n = len(chunks)
    cc = [c.char_count for c in chunks]
    tc = [c.est_token_count for c in chunks]
    n_table = sum(1 for c in chunks if c.has_table)
    n_split = sum(1 for c in chunks if c.is_split_part)
    n_tiny = sum(1 for c in chunks if c.char_count < min_chars)
    source = chunks[0].source
    company = chunks[0].company

    lines = []
    lines.append(f"# 切分输出: {source}")
    if company:
        lines.append(f"**公司**: {company}")
    lines.append("")
    lines.append("> 本文档供切分质量审阅。每个 `##` 标题是一个 chunk，内容按原 markdown 渲染。")
    lines.append("> 若发现表格被切残 / 引导句被吸到上一块 / 块之间语义断裂，请反馈以调整切分参数。")
    lines.append("")

    # 总览统计
    lines.append("## 总览统计")
    lines.append("")
    lines.append(f"- **总块数**: {n}")
    if cc:
        lines.append(f"- **字符数** (min / avg / max): {min(cc)} / {sum(cc) // n} / {max(cc)}")
        lines.append(f"- **估算 token** (min / avg / max): {min(tc)} / {sum(tc) // n} / {max(tc)}")
    lines.append(f"- **含表格的块数**: {n_table} ({n_table * 100 // n}%  )")
    lines.append(f"- **被进一步切开的子块数**: {n_split}")
    lines.append(f"- **过短 (<{min_chars} 字符) 的块数**: {n_tiny}  ← 仅告警")
    # chunk_type 分布
    from collections import Counter
    type_counts = Counter(c.chunk_type for c in chunks)
    lines.append(f"- **chunk_type 分布**: {dict(type_counts)}")
    lines.append("")

    # 目录
    lines.append("## 目录")
    lines.append("")
    for i, c in enumerate(chunks, start=1):
        bc = " > ".join(c.breadcrumb) if c.breadcrumb else "(前言)"
        ct = f"[{c.chunk_type}]"
        flag = " [表]" if c.has_table else ""
        split_note = f" *(子块 {c.split_part_index}/{c.split_part_total})*" if c.is_split_part else ""
        lines.append(f"{i}. `{c.chunk_id}` {ct}{flag} — {bc}{split_note}")
    lines.append("")

    # 逐块内容
    for i, c in enumerate(chunks, start=1):
        bc = " > ".join(c.breadcrumb) if c.breadcrumb else "(前言)"
        flag = " [表]" if c.has_table else ""
        anchor = c.chunk_id.lower().replace("_", "-")

        lines.append("---")
        lines.append("")
        lines.append(f'<a id="{anchor}"></a>')
        lines.append(f"## 块 {i} / {n}: `{c.chunk_id}` {flag}")
        lines.append("")

        meta_bits = [
            f"**类型**: {c.chunk_type}",
            f"**路径**: {bc}",
            f"**章节**: {c.section_title or '(无)'} (H{c.heading_level})",
            f"**{c.char_count} 字符 ≈ {c.est_token_count} tokens**",
        ]
        if c.parent_id:
            meta_bits.append(f"**父块**: {c.parent_id}")
        if c.children_ids:
            meta_bits.append(f"**子块**: {len(c.children_ids)} 个")
        if c.sibling_ids:
            meta_bits.append(f"**兄弟块**: {len(c.sibling_ids)} 个")
        if c.is_split_part:
            meta_bits.append(f"**子块**: {c.split_part_index} / {c.split_part_total}")
        if c.table_caption:
            meta_bits.append(f"**表格**: {c.table_caption}")
        lines.append(" · ".join(meta_bits))
        lines.append("")
        lines.append(c.content.rstrip())
        lines.append("")

    with open(path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))


def print_summary(chunks) -> None:
    """打印分块摘要统计。"""
    if not chunks:
        print("  [chunker] 0 chunks")
        return
    n = len(chunks)
    cc = [c.char_count for c in chunks]
    tc = [c.est_token_count for c in chunks]
    type_counts = Counter(c.chunk_type for c in chunks)
    print(f"  [chunker] 总块数: {n}")
    print(f"    类型分布: {dict(type_counts)}")
    print(f"    字符 min/avg/max: {min(cc)} / {sum(cc)//n} / {max(cc)}")
    print(f"    Token min/avg/max: {min(tc)} / {sum(tc)//n} / {max(tc)}")
