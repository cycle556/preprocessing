"""

策略（三层 + 表格感知）:
    1) 按 ## (H2) 主章节切粗块
    2) 超过 MAX_CHARS 的块按 ### / #### 继续往下切
    3) 仍超大的叶子块再细切:
        - 含 Markdown 表格 -> 按行切, 每个子块复制表头与分隔行(自包含)
        - 纯散文        -> 中文感知递归字符切分(带 overlap)
    4) 每个块前回填面包屑(标题路径)作为上下文头, 提升检索召回与下游 LLM 阅读体验

输出: JSONL / XLSX / MD 三种格式
    - JSONL: 进 RAG / embedding 流水线
    - XLSX : 标注 (一行一块, 右侧加 question/answer 列)
    - MD   : 审阅切分质量, 表格能真实渲染

依赖: 仅标准库; tiktoken 与 openpyxl 可选

特殊：SOURCE_NAME = 改
读到287
"""


import json
import re
import os
from dataclasses import dataclass, field, asdict
from typing import List, Tuple, Optional

# openpyxl用于输出 xlsx
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

#-------------------
#1. 配置
#-------------------

# 块大小目标
MAX_CHARS = 900  # 单块字符上限，超过则触发细切
MIN_CHARS = 80  # 单块字符下限，仅用于日志告警，不做强制合并(短小章节视作一个语义单元)
OVERLAP_CHARS = 80  # 散文递归切分的重叠字符数，表格切分不做 overlap (按行独立)

# 文档来源标识 (会写入每条 chunk 的 source 字段, 后续可用于多文档 RAG 的过滤)
SOURCE_NAME = "Chubb_Manual 高資產業務人壽保障產品指引 (CH)"

# 中文递归切分分隔符, 从粗到细
RECURSIVE_SEPARATORS = ["\n\n", "\n", "。", "！", "？", "；", "，", " ", ""]

#------------------
# 2. 数据结构
#------------------

@dataclass
class Chunk:
    chunk_id: str
    source: str
    breadcrumb: List[str]  # 标题路径
    section_title: str  # 当前块所属最深一级的标题
    heading_level: int  # 0 = 前言/无标题, 1 = #, 2 = ##, ...
    content: str  # 原始正文(不含面包屑回填)
    content_with_context: str  # 回填面包屑后的内容, 推荐用于 embedding & LLM 输入
    char_count: int
    est_token_count: int
    has_table: bool
    is_split_part: bool = False  # True 表示该块是被进一步细切出的子块
    split_part_index: Optional[int] = None  # 子块序号(从 1 开始)
    split_part_total: Optional[int] = None  # 同一父块的子块总数


#---------------------------------------
# 3. Token 估算 (可选 tiktoken)
#---------------------------------------

def _make_token_counter():
    """优先使用 tiktoken cl100k_base; 没有则回退到字符级启发式."""
    try:
        import tiktoken
        enc = tiktoken.get_encoding("cl100k_base")
        return lambda s: len(enc.encode(s))
    except Exception:
        # 中文每字 ≈ 1.5 token，英文按 4 char/token 近似，粗略计算
        def heuristic(s: str) -> int:
            cjk = sum(1 for ch in s if '\u4e00' <= ch <= '\u9fff')
            other = len(s) - cjk
            return int(cjk * 1.5 + other / 4)

        return heuristic


count_tokens = _make_token_counter()

#--------------------------------------------------------
# 4. Markdown 结构解析: 提取 (level, heading, body)
#--------------------------------------------------------

HEADING_RE = re.compile(r'^(#{1,6})\s+(.+?)\s*$') #提取一行里的标题等级和标题内容
CODE_FENCE_RE = re.compile(r'^\s*```') #识别一行是否是markdown代码块的开始 / 结束标记。


def normalize_markdown_tables(md_text: str) -> str:
    """折叠表格行单元格内的多余空格.

    很多 md 文档(尤其是从 PDF / docx 转出来的)会把每个单元格用空格补齐到统一宽度,
    导致一行普通 '| ME | 身體檢查 |' 也会变成 300+ 字符. 这些填充空格对 RAG 没有任何
    语义价值, 却严重干扰字符数统计与切分判断, 需要统一压缩.

    只对识别为 '表格行' 的行做处理(以 '|' 开头/结尾, 或与表格分隔行相邻),
    其它行原样保留.
    """
    lines = md_text.split('\n')
    # 标记每行是否是表格行 (基于结构: 前/后一行是分隔行, 或与分隔行同段)
    is_table_line = [False] * len(lines)
    #实现整块 Markdown 表格的识别。
    for i, ln in enumerate(lines):
        if _is_table_separator(ln):
            is_table_line[i] = True  #判断是否为表格分隔符----
            # 向上向下扩散
            j = i - 1
            while j >= 0 and '|' in lines[j] and lines[j].strip():
                is_table_line[j] = True
                j -= 1
            j = i + 1
            while j < len(lines) and '|' in lines[j] and lines[j].strip():
                is_table_line[j] = True
                j += 1

    #标准化：分隔线统一为最短横线、内容行压缩多余空格，非表格行不动，输出格式整齐的 Markdown 表格文本。
    out = []
    for ln, is_tbl in zip(lines, is_table_line):
        if is_tbl:
            if _is_table_separator(ln):
                # 分隔行: 把每个 cell 内的 -/: 压缩到最短表达 (不影响对齐语义)
                cells = ln.split('|')
                norm = []
                for c in cells:
                    s = c.strip()
                    if not s:
                        continue
                    # 保留左/右对齐冒号, 但把 --- 减到 3 个
                    left = ':' if s.startswith(':') else ''
                    right = ':' if s.endswith(':') else ''
                    norm.append(f'{left}---{right}')
                out.append('| ' + ' | '.join(norm) + ' |')
            else:
                # 普通表格行: 折叠 cell 内的连续空白
                cells = ln.split('|')
                cells = [re.sub(r'[ \t]+', ' ', c).strip() for c in cells]
                out.append('| ' + ' | '.join(c for c in cells if c != '') + ' |')
        else:
            out.append(ln)
    return '\n'.join(out)


def parse_sections(md_text: str) -> List[Tuple[int, str, str]]:
    """把 Markdown 拆成 (heading_level, heading_text, body_text) 列表.

    level=0 表示首个标题之前的前言/无标题内容(若存在).
    会跳过代码围栏内部的 # 行, 避免被误判为标题.
    """
    sections: List[Tuple[int, str, str]] = []
    lines = md_text.splitlines()

    cur_level, cur_heading, cur_body = 0, "", []
    in_code = False

    #解析 Markdown 文本，提取各级标题 + 对应正文段落，并忽略代码块内内容，最终存入 sections 列表。
    for line in lines:
        if CODE_FENCE_RE.match(line):
            in_code = not in_code
            cur_body.append(line)
            continue

        if not in_code:
            m = HEADING_RE.match(line)
            if m:
                # 切换章节: 先收尾上一段
                body_text = "\n".join(cur_body).strip("\n")
                if cur_level != 0 or body_text:
                    sections.append((cur_level, cur_heading, body_text))
                cur_level = len(m.group(1))
                cur_heading = m.group(2).strip()
                cur_body = []
                continue

        cur_body.append(line)

    # flush
    #刷新缓存，将有效章节数据保存下来，过滤掉无层级且无正文的空段
    body_text = "\n".join(cur_body).strip("\n")
    if cur_level != 0 or body_text:
        sections.append((cur_level, cur_heading, body_text)) #把「层级 + 标题 + 正文」三元组加入章节列表，完成一个章节的缓存落盘（flush 刷新）。

    return sections


def build_breadcrumbs(sections: List[Tuple[int, str, str]]
                      ) -> List[Tuple[List[str], int, str, str]]:
    """为每个 section 构建面包屑路径.

    返回 (breadcrumb_path, heading_level, heading_text, body) 列表.
    """
    out = []
    stack: List[Tuple[int, str]] = []  # (level, heading)

    for level, heading, body in sections:
        if level == 0:
            out.append(([], 0, "", body))
            continue
        while stack and stack[-1][0] >= level:
            stack.pop()
        stack.append((level, heading))
        breadcrumb = [h for _, h in stack]
        out.append((breadcrumb, level, heading, body))

    return out

#-------------------------------
# 5. 表格检测与按行切分
#-------------------------------
TABLE_SEP_RE = re.compile(r'^\s*\|?\s*:?-{2,}.*\|.*-{2,}', re.MULTILINE)


def _is_table_separator(line: str) -> bool:
    s = line.strip().strip('|').strip()
    if not s:
        return False
    parts = [p.strip() for p in s.split('|')]
    # 至少 2 列, 且每列都形如 :--- / :---: / ---:
    return len(parts) >= 2 and all(re.fullmatch(r':?-{2,}:?', p) for p in parts if p)


def split_text_into_blocks(text: str) -> List[Tuple[str, str]]:
    """把 section body 切成 [(kind, content), ...] 列表; kind ∈ {'table', 'prose'}."""
    lines = text.split('\n')
    blocks: List[Tuple[str, str]] = []
    i, n = 0, len(lines)

    while i < n:
        # 表格起点: 当前行含 '|' 且下一行是分隔行
        if '|' in lines[i] and i + 1 < n and _is_table_separator(lines[i + 1]):
            j = i + 2
            while j < n and '|' in lines[j] and lines[j].strip():
                j += 1
            blocks.append(('table', '\n'.join(lines[i:j])))
            i = j
        else:
            j = i
            while j < n:
                if (j + 1 < n and '|' in lines[j]
                        and _is_table_separator(lines[j + 1])):
                    break
                j += 1
            prose = '\n'.join(lines[i:j]).strip('\n')
            if prose:
                blocks.append(('prose', prose))
            i = j

    return blocks


def split_table_by_rows(table_text: str, max_chars: int) -> List[str]:
    """将文本格式的表格按字符长度上限切分，每个子表都保留表头 + 分隔行，避免拆分后丢失表头"""
    lines = [ln for ln in table_text.split('\n') if ln.strip()]
    if len(lines) < 3:
        return [table_text]

    header, sep = lines[0], lines[1]
    data_rows = lines[2:]
    head_size = len(header) + len(sep) + 2

    out: List[str] = []
    cur, cur_size = [header, sep], head_size
    for row in data_rows:
        row_size = len(row) + 1
        # 即使单行就超长, 也要让它独立成块 —— 否则会出现 0 数据行的空块
        if cur_size + row_size > max_chars and len(cur) > 2:
            out.append('\n'.join(cur))
            cur, cur_size = [header, sep], head_size
        cur.append(row)
        cur_size += row_size

    if len(cur) > 2:
        out.append('\n'.join(cur))

    return out

#--------------------------------
# 6. 中文感知递归字符切分 (散文用)
#--------------------------------

def recursive_char_split(text: str, max_chars: int, overlap: int,
                         seps: List[str] = RECURSIVE_SEPARATORS) -> List[str]:
    """从粗到细尝试分隔符, 找到一个能把 text 切到 <= max_chars 的, 然后合并 + overlap."""
    if len(text) <= max_chars:
        return [text]

    for sep in seps:
        if sep == "":
            # 兜底: 按字符硬切
            return _hard_char_split(text, max_chars, overlap)
        if sep not in text:
            continue
        pieces = _split_with_sep(text, sep)
        # 若拆出来还有片段 > max_chars, 对这些片段继续递归用下一层 sep
        refined: List[str] = []
        for p in pieces:
            if len(p) > max_chars:
                refined.extend(recursive_char_split(p, max_chars, overlap,
                                                    seps[seps.index(sep) + 1:]))
            else:
                refined.append(p)
        return _merge_with_overlap(refined, max_chars, overlap)

    return _hard_char_split(text, max_chars, overlap)


def _split_with_sep(text: str, sep: str) -> List[str]:
    """按 sep 切, 保留 sep 在前一片段尾部, 避免标点丢失."""
    if sep == "\n\n" or sep == "\n":
        parts = text.split(sep)
        return [p + (sep if i < len(parts) - 1 else "") for i, p in enumerate(parts) if p != ""]
    out, buf = [], ""
    for ch in text:
        buf += ch
        if ch == sep:
            out.append(buf)
            buf = ""
    if buf:
        out.append(buf)
    return out


def _merge_with_overlap(pieces: List[str], max_chars: int, overlap: int) -> List[str]:
    """把小片段贪心合并到 max_chars, 块与块之间留 overlap 字符."""
    out: List[str] = []
    cur = ""
    for p in pieces:
        if not cur:
            cur = p
        elif len(cur) + len(p) <= max_chars:
            cur += p
        else:
            out.append(cur)
            tail = cur[-overlap:] if overlap and len(cur) > overlap else ""
            cur = tail + p
            # 如果 tail + p 本身就超了, 不再二次切 —— 上层已确保 p<=max_chars
    if cur:
        out.append(cur)
    return out


def _hard_char_split(text: str, max_chars: int, overlap: int) -> List[str]:
    out, i = [], 0
    while i < len(text):
        out.append(text[i:i + max_chars])
        i += max(1, max_chars - overlap)
    return out


# ----------------------------------------------------------
# 7. 主调度: 把单个 section 切成 N 个 chunk content
# ----------------------------------------------------------

def _looks_like_table_caption(text: str) -> bool:
    """判断一段短小 prose 是否像紧跟其后那张表格的"标题/引导句".

    例: 一张表上方有一行没有用 markdown 标题语法标记的小标题, 例如:
        '兒童（0-17 歲）驗身要求（適用於RSL）'
        | 總保障額 | 年齡 0-17 |
        | --- | --- |
        ...
    这一行从语义上是后面那张表的标题, 应该和它一起被切到同一个 chunk 里,
    而不是被它前面那张表"吸走"成为前一个 chunk 的尾巴.

    判断方法：
      a) 末尾是冒号 (引出下文的强信号), 不限长度;
      b) 行数 <=2 且总长度 <=80 字符且末尾不是句子终止符;
    """
    s = text.strip()
    if not s:
        return False
    # 末尾冒号 -> 引出下文
    if s.rstrip().endswith((':', '：')):
        return True
    lines = [ln for ln in s.split('\n') if ln.strip()]
    if len(lines) > 2:
        return False
    if len(s) > 80:
        return False
    last_char = s.rstrip()[-1]
    # 完整句子的标志: 各种句末标点
    if last_char in '。！？；.!?':
        return False
    return True


def chunk_section_body(body: str, max_chars: int, overlap: int) -> List[Tuple[str, bool]]:
    """对 section body 进行细切, 返回 [(content, has_table), ...].

    策略:
        1) 先把 body 切成 prose / table 块序列
        2) 任何超大的单块再细切(表格按行 / 散文递归)
        3) 贪心合并相邻小块, 累计到 max_chars 才切开 ——
           关键: 这能避免短小引导句脱离它紧跟着的表格而独立成块.
    """
    if not body.strip():
        return []
    if len(body) <= max_chars:
        return [(body, _contains_table(body))]

    # --- 第一步: 切块 + 对超大块二次细切 ---
    expanded: List[Tuple[str, str]] = []  # [(kind, content)]
    for kind, blk in split_text_into_blocks(body):
        if len(blk) <= max_chars:
            expanded.append((kind, blk))
        elif kind == 'table':
            for sub in split_table_by_rows(blk, max_chars):
                expanded.append(('table', sub))
        else:
            for sub in recursive_char_split(blk, max_chars, overlap):
                expanded.append(('prose', sub))

    # --- 第二步: 贪心合并相邻块 ---
    # 关键修复(前瞻一步): 若当前块是"短小、像表格标题的 prose", 且下一块是 table,
    # 则它语义上属于下一张表, 不应被前面一张表吸走 —— 此时若 cur 非空, 先 flush.
    out: List[Tuple[str, bool]] = []
    cur_parts: List[str] = []
    cur_size = 0
    cur_has_table = False
    SEP = "\n\n"
    n_expanded = len(expanded)

    for idx, (kind, blk) in enumerate(expanded):
        add_size = len(blk) + (len(SEP) if cur_parts else 0)

        is_table_caption = (
                kind == 'prose'
                and idx + 1 < n_expanded
                and expanded[idx + 1][0] == 'table'
                and _looks_like_table_caption(blk)
        )

        # 触发 flush 的两种情形:
        #   1) 加入新块会超过 max_chars (原有逻辑)
        #   2) 当前块是表格引导句, 应该和下一张表绑在一起, 而 cur 已经有内容了
        need_flush = cur_parts and (
                (cur_size + add_size > max_chars)
                or is_table_caption
        )
        if need_flush:
            out.append((SEP.join(cur_parts), cur_has_table))
            cur_parts = []
            cur_size = 0
            cur_has_table = False
            add_size = len(blk)

        cur_parts.append(blk)
        cur_size += add_size
        if kind == 'table':
            cur_has_table = True

    if cur_parts:
        out.append((SEP.join(cur_parts), cur_has_table))

    return out


def _contains_table(text: str) -> bool:
    lines = text.split('\n')
    for i in range(len(lines) - 1):
        if '|' in lines[i] and _is_table_separator(lines[i + 1]):
            return True
    return False


# ---------------------------------------
# 8. 顶层入口: 文本 -> [Chunk]
# ---------------------------------------

def chunk_markdown(md_text: str,
                   source: str = SOURCE_NAME,
                   max_chars: int = MAX_CHARS,
                   overlap: int = OVERLAP_CHARS,
                   id_prefix: str = "chunk") -> List[Chunk]:
    """完整切分管线."""
    # 预处理: 折叠表格单元格内的对齐填充空格(只影响字符数统计, 不影响语义)
    md_text = normalize_markdown_tables(md_text)

    sections = parse_sections(md_text)
    walked = build_breadcrumbs(sections)

    # 第一遍: 把 (breadcrumb, body) 按章节级别合并 —— 如果一个 ## 章节整体小于 max_chars,
    # 它的子标题 ### / #### 也合并进来, 避免出现"父章节没正文 + 子章节散块"的零碎现象.
    merged: List[Tuple[List[str], int, str, str]] = _merge_small_subsections(
        walked, max_chars)

    # 第二遍: 对每个合并后的单元做细切
    chunks: List[Chunk] = []
    counter = 0
    for breadcrumb, level, heading, body in merged:
        if not body.strip():
            continue
        parts = chunk_section_body(body, max_chars, overlap)
        total = len(parts)
        for idx, (content, has_table) in enumerate(parts, start=1):
            counter += 1
            ctx_header = " > ".join(breadcrumb) if breadcrumb else "(前言)"
            content_with_context = f"# {ctx_header}\n\n{content}"
            chunks.append(Chunk(
                chunk_id=f"{id_prefix}_{counter:04d}",
                source=source,
                breadcrumb=breadcrumb,
                section_title=heading,
                heading_level=level,
                content=content,
                content_with_context=content_with_context,
                char_count=len(content),
                est_token_count=count_tokens(content_with_context),
                has_table=has_table,
                is_split_part=(total > 1),
                split_part_index=idx if total > 1 else None,
                split_part_total=total if total > 1 else None,
            ))
    return chunks


def _merge_small_subsections(walked, max_chars: int):
    """若某 ## 章节总字符 <= max_chars, 把其全部子节合并为一个单元
    (合并后 body 会带上 ### / #### 标题原文, 保留层级语义).

    否则按原结构展开 ## / ### / #### 各自独立.
    """
    # 1) 先索引: 找出每个 H2 段的开始和结束
    h2_indices = [i for i, (_, lvl, *_) in enumerate(walked) if lvl == 2]
    if not walked:
        return []

    out = []
    # 先处理"第一个 H2 之前的所有 section": 包括 level=0 的前言, 也包括 H1
    # 以及任何直接挂在 H1 下、没有被 H2 包裹的 H3/H4 (虽然少见, 但保留以避免内容丢失).
    first_h2 = h2_indices[0] if h2_indices else len(walked)
    for item in walked[:first_h2]:
        # 跳过空 body 的纯标题占位
        if item[1] == 0 and not item[3].strip():
            continue
        out.append(item)

    for k, start in enumerate(h2_indices):
        end = h2_indices[k + 1] if k + 1 < len(h2_indices) else len(walked)
        group = walked[start:end]

        # 计算合并后的总大小
        h2_breadcrumb, h2_level, h2_heading, h2_body = group[0]
        merged_body_parts = []
        if h2_body.strip():
            merged_body_parts.append(h2_body)
        for sub in group[1:]:
            sub_bc, sub_lvl, sub_h, sub_b = sub
            hashes = "#" * sub_lvl
            merged_body_parts.append(f"{hashes} {sub_h}\n{sub_b}".rstrip())
        merged_body = "\n\n".join(merged_body_parts).strip()

        if len(merged_body) <= max_chars or len(group) == 1:
            # 整个 ## 章节足够小 -> 合并为一个单元
            out.append((h2_breadcrumb, h2_level, h2_heading, merged_body))
        else:
            # 太大 -> 展开为各自独立单元(H2 自身正文 + 每个子节)
            if h2_body.strip():
                out.append((h2_breadcrumb, h2_level, h2_heading, h2_body))
            else:
                # H2 标题下没有直接正文; 仍保留一个标题占位以维持面包屑(可选, 这里跳过空块)
                pass
            for sub in group[1:]:
                out.append(sub)

    return out


# -------------------------------------
# 9. CLI / 主入口
# -------------------------------------

def write_jsonl(chunks: List[Chunk], path: str) -> None:
    with open(path, "w", encoding="utf-8") as f:
        for c in chunks:
            f.write(json.dumps(asdict(c), ensure_ascii=False) + "\n")


def write_xlsx(chunks: List[Chunk], path: str) -> None:
    """把 chunks 写入 xlsx, 一行一个块, 表头冻结 + 筛选器 + 文本换行.

    用途: 方便在 Excel 中浏览块、按 breadcrumb/has_table 筛选
    """
    if not HAS_OPENPYXL:
        raise ImportError("写入 xlsx 需要先安装 openpyxl: pip install openpyxl")

    wb = Workbook()
    ws = wb.active
    ws.title = "chunks"

    headers = [
        "chunk_id", "source", "breadcrumb", "section_title", "heading_level",
        "char_count", "est_token_count", "has_table",
        "is_split_part", "split_part_index", "split_part_total",
        "content", "content_with_context",
    ]
    ws.append(headers)

    for c in chunks:
        ws.append([
            c.chunk_id,
            c.source,
            " > ".join(c.breadcrumb),
            c.section_title,
            c.heading_level,
            c.char_count,
            c.est_token_count,
            c.has_table,
            c.is_split_part,
            c.split_part_index if c.split_part_index is not None else "",
            c.split_part_total if c.split_part_total is not None else "",
            c.content,
            c.content_with_context,
        ])

    # --- 样式: 表头 ---
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", start_color="305496")
    header_align = Alignment(horizontal="center", vertical="center")
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # --- 列宽 ---
    widths = {
        "chunk_id": 16, "source": 30, "breadcrumb": 50, "section_title": 26,
        "heading_level": 8, "char_count": 9, "est_token_count": 10,
        "has_table": 9, "is_split_part": 10,
        "split_part_index": 9, "split_part_total": 9,
        "content": 60, "content_with_context": 70,
    }
    for i, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(h, 15)

    # --- 数据行: 普通字段顶端对齐, content 类字段开换行 ---
    wrap_cols = {headers.index(h) + 1 for h in
                 ("breadcrumb", "content", "content_with_context")}
    body_font = Font(name="Arial", size=10)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = body_font
            if cell.column in wrap_cols:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            else:
                cell.alignment = Alignment(vertical="top")

    # --- 冻结表头 + 自动筛选 ---
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(path)


def _md_anchor(chunk_id: str) -> str:
    """生成在 markdown 渲染器中可用的锚点 id (小写, 简单字符)."""
    return chunk_id.lower().replace("_", "-")


def write_md(chunks: List[Chunk], path: str) -> None:
    """把 chunks 写入 markdown 审阅文档.

    用途: 做切分质量审阅. 相比 xlsx, 核心优势是:
        - 表格能被 markdown 渲染器真实渲染出来
        - 上下滚动天然连续, 便于判断"相邻块语义是否被切断"
        - 锚点目录可一键跳转到可疑块

    结构:
        1. 文件头: 文档名 + 总览统计
        2. 目录: 锚点跳转到每个块, 含表格的块加 📊 标记
        3. 每块一段:
            - ## 块 i/N: chunk_id  (含表格的话加 📊)
            - 紧凑 metadata 行 (面包屑 / 字符 / token / 子块编号)
            - 内容原文 (markdown 原样写入, 表格会被渲染)
    """
    if not chunks:
        with open(path, "w", encoding="utf-8") as f:
            f.write("# 切分输出: (空)\n\n没有产生任何块.\n")
        return

    n = len(chunks)
    cc = [c.char_count for c in chunks]
    tc = [c.est_token_count for c in chunks]
    n_table = sum(1 for c in chunks if c.has_table)
    n_split = sum(1 for c in chunks if c.is_split_part)
    n_tiny = sum(1 for c in chunks if c.char_count < MIN_CHARS)
    source = chunks[0].source

    lines: List[str] = []

    # --- 文件头 ---
    lines.append(f"# 切分输出: {source}")
    lines.append("")
    lines.append("> 块内有表格的话，就显示📊")
    #lines.append("> 本文档供切分质量审阅. 每个 `##` 标题是一个 chunk, 内容按原 markdown 渲染.")
    #lines.append("> 若发现表格被切残 / 引导句被吸到上一块 / 块之间语义断裂等问题, 请反馈以调整切分参数.")
    lines.append("")

    # --- 总览统计 ---
    # lines.append("## 总览统计")
    # lines.append("")
    # lines.append(f"- **总块数**: {n}")
    # lines.append(f"- **字符数** (min / avg / max): {min(cc)} / {sum(cc) // n} / {max(cc)}")
    # lines.append(f"- **估算 token** (min / avg / max): {min(tc)} / {sum(tc) // n} / {max(tc)}")
    # lines.append(f"- **含表格的块数**: {n_table} ({n_table * 100 // n}%)")
    # lines.append(f"- **被进一步切开的子块数**: {n_split}")
    # lines.append(f"- **过短 (<{MIN_CHARS} 字符) 的块数**: {n_tiny}  ← 仅告警, 不强制合并")
    # lines.append("")

    # --- 目录 ---
    lines.append("## 目录")
    lines.append("")
    for i, c in enumerate(chunks, start=1):
        bc = " > ".join(c.breadcrumb) if c.breadcrumb else "(前言)"
        flag = " 📊" if c.has_table else ""
        split_note = ""
        if c.is_split_part:
            split_note = f" *(子块 {c.split_part_index}/{c.split_part_total})*"
        anchor = _md_anchor(c.chunk_id)
        lines.append(f"{i}. [`{c.chunk_id}`](#{anchor}){flag} — {bc}{split_note}")
    lines.append("")

    # --- 逐块 ---
    for i, c in enumerate(chunks, start=1):
        bc = " > ".join(c.breadcrumb) if c.breadcrumb else "(前言)"
        flag = " 📊" if c.has_table else ""
        anchor = _md_anchor(c.chunk_id)

        lines.append("---")
        lines.append("")
        # HTML 锚点保证在 GitHub / Obsidian / VSCode preview 都能跳转
        lines.append(f'<a id="{anchor}"></a>')
        lines.append(f"## 块 {i} / {n}: `{c.chunk_id}`{flag}")
        lines.append("")

        # 紧凑 metadata: 一行能塞下的全塞一行, 用 · 分隔
        meta_bits = [
            f"**路径**: {bc}",
            f"**章节**: {c.section_title or '(无)'} (H{c.heading_level})",
            # f"**{c.char_count} 字符 ≈ {c.est_token_count} tokens**",
        ]
        if c.is_split_part:
            meta_bits.append(f"**子块**: {c.split_part_index} / {c.split_part_total}")
        if c.has_table:
            meta_bits.append("**含表格**")
        lines.append(" · ".join(meta_bits))
        lines.append("")

        # 内容原文不要包代码块, 否则 markdown 表格无法渲染.
        # content 内部可能含 ### / #### 子标题 (来自 _merge_small_subsections 的回填),
        # 它们在视觉上会出现在我们的 ## 块标题之下, 层级看起来是合理的.
        lines.append(c.content.rstrip())
        lines.append("")

    with open(path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))


def print_summary(chunks: List[Chunk]) -> None:
    n = len(chunks)
    if n == 0:
        print("[WARN] 0 chunks produced.")
        return
    cc = [c.char_count for c in chunks]
    tc = [c.est_token_count for c in chunks]
    n_table = sum(1 for c in chunks if c.has_table)
    n_split = sum(1 for c in chunks if c.is_split_part)
    n_tiny = sum(1 for c in chunks if c.char_count < MIN_CHARS)
    print(f"总块数: {n}")
    print(f"字符数 - min/avg/max: {min(cc)} / {sum(cc) // n} / {max(cc)}")
    print(f"估算token - min/avg/max: {min(tc)} / {sum(tc) // n} / {max(tc)}")
    print(f"含表格的块数: {n_table}  ({n_table * 100 // n}% 占比)")
    print(f"被进一步切开的子块数: {n_split}")
    print(f"过短(<{MIN_CHARS}字符)的块数: {n_tiny}  <- 仅告警, 视为短小语义单元保留")


def main(input_path: str, output_path: str) -> None:
    with open(input_path, "r", encoding="utf-8") as f:
        md_text = f.read()

    chunks = chunk_markdown(md_text, source=SOURCE_NAME,
                            max_chars=MAX_CHARS, overlap=OVERLAP_CHARS,
                            id_prefix="chunk")

    # 同时输出 JSONL / XLSX / MD
    base, _ = os.path.splitext(output_path)
    jsonl_path = base + ".jsonl"
    xlsx_path = base + ".xlsx"
    md_path = base + ".md"

    write_jsonl(chunks, jsonl_path)
    print(f"[OK] JSONL 写入 {jsonl_path}")

    if HAS_OPENPYXL:
        write_xlsx(chunks, xlsx_path)
        print(f"[OK] XLSX  写入 {xlsx_path}")
    else:
        print("[SKIP] XLSX 未生成 -- 请先 'pip install openpyxl' 后重新运行")

    write_md(chunks, md_path)
    print(f"[OK] MD    写入 {md_path}")

    print_summary(chunks)


if __name__ == "__main__":
    import sys

    if len(sys.argv) >= 3:
        main(sys.argv[1], sys.argv[2])
    else:
        default_in = "Chubb_Manual 高資產業務人壽保障產品指引 (CH) - Clean_2024Oct.md"
        default_out = "chunks.jsonl"
        main(default_in, default_out)