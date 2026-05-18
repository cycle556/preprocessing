# -*- coding: utf-8 -*-
"""
这是一个通用的 Markdown 文本切分器（支持中日韩文字 CJK + 英文，且能识别表格）。

策略：
    1) 按主要章节层级进行切分（自动检测，不再硬编码为 H2）。
    2) 如果每个主要章节下的子章节总大小在 MAX_CHARS 限制内，则将其合并，
       以保持较小的章节内容完整。
    3) 对于过大的末级区块，将进一步切分：
        Markdown 表格 -> 按数据行切分，每个分片都复制表头和分隔行。
        散文/段落文本 -> 使用支持中日韩文字和英文的分隔符进行递归字符切分。
    4) 将“面包屑”（标题路径）回填到每个文本块中，以供检索和提供给大语言模型（LLM）上下文。

输出格式：JSONL / XLSX / MD（可通过 --formats 参数选择任意子集）。

主要的通用化改进点：
    主要标题层级从文档中自动检测（不再硬编码为 H2）。
    源文件名和块 ID 前缀均从输入文件名派生。
    递归切分的分隔符包含了英文句子结束符（". ", "! ", "? ", ...）。
    _split_with_sep 函数支持多字符分隔符（以前仅支持单字符）。
    基于 argparse 的命令行接口（CLI），支持批量处理 / 通配符匹配，以及每次运行时覆盖参数。
"""

from __future__ import annotations

import argparse
import glob
import json
import os
import re
import sys
from collections import Counter
from dataclasses import dataclass, asdict
from typing import List, Tuple, Optional


try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# ============================================================================
# 1. 默认配置
# ============================================================================

DEFAULT_MAX_CHARS = 1800
DEFAULT_MIN_CHARS = 80      
DEFAULT_OVERLAP_CHARS = 80

# 递归切分器的分隔符按“从粗到细”的顺序尝试。同时兼容中日韩文字和英文，
#让切分器在处理这两种语言时都能有效切分，而不至于退化到逐字符切分。
RECURSIVE_SEPARATORS = [
    "\n\n", "\n",
    "。", "！", "？",          # 句尾标记
    ". ", "! ", "? ",         # ASCII句末结束符（句点+空格；比单独使用“.”更稳妥）
    "；", "; ",
    "，", ", ",
    " ", "",
]


# ============================================================================
# 2. 数据结构
# ============================================================================

@dataclass
class Chunk:
    chunk_id: str
    source: str
    breadcrumb: List[str]
    section_title: str
    heading_level: int
    content: str
    content_with_context: str
    char_count: int
    est_token_count: int
    has_table: bool
    table_caption: Optional[str] = None
    is_split_part: bool = False
    split_part_index: Optional[int] = None
    split_part_total: Optional[int] = None


# ============================================================================
# 3. 3. Token 估算 (可选 tiktoken)
# ============================================================================

def _make_token_counter():
    try:
        import tiktoken
        enc = tiktoken.get_encoding("cl100k_base")
        return lambda s: len(enc.encode(s))
    except Exception:
        def heuristic(s: str) -> int:
            cjk = sum(1 for ch in s if '\u4e00' <= ch <= '\u9fff')
            other = len(s) - cjk
            return int(cjk * 1.5 + other / 4)
        return heuristic


count_tokens = _make_token_counter()


# ============================================================================
# 4. Markdown 结构解析
# ============================================================================

HEADING_RE = re.compile(r'^(#{1,6})\s+(.+?)\s*$')
CODE_FENCE_RE = re.compile(r'^\s*```')


def normalize_markdown_tables(md_text: str) -> str:
    """压缩 Markdown 表格单元格内部多余的空白字符。
       
       从 PDF 或 Word 文档转换而来的表格，其单元格往往会被填充大量空格来保持统一的视觉宽度；
       这些空格在语义上是毫无意义的，但却会虚增字符数量，进而干扰切分时的长度判断。
       因此，我们需要把这些多余空格压缩掉，同时保证不影响非表格内容的原始排版。
    """
    lines = md_text.split('\n')
    is_table_line = [False] * len(lines)
    for i, ln in enumerate(lines):
        if _is_table_separator(ln):
            is_table_line[i] = True
            j = i - 1
            while j >= 0 and '|' in lines[j] and lines[j].strip():
                is_table_line[j] = True
                j -= 1
            j = i + 1
            while j < len(lines) and '|' in lines[j] and lines[j].strip():
                is_table_line[j] = True
                j += 1

    out = []
    for ln, is_tbl in zip(lines, is_table_line):
        if is_tbl:
            if _is_table_separator(ln):
                cells = ln.split('|')
                norm = []
                for c in cells:
                    s = c.strip()
                    if not s:
                        continue
                    left = ':' if s.startswith(':') else ''
                    right = ':' if s.endswith(':') else ''
                    norm.append(f'{left}---{right}')
                out.append('| ' + ' | '.join(norm) + ' |')
            else:
                cells = ln.split('|')
                cells = [re.sub(r'[ \t]+', ' ', c).strip() for c in cells]
                out.append('| ' + ' | '.join(c for c in cells if c != '') + ' |')
        else:
            out.append(ln)
    return '\n'.join(out)


def parse_sections(md_text: str) -> List[Tuple[int, str, str]]:
    """
    将Markdown拆分为（标题级别、标题文本、正文）三元组。
    级别0代表前言部分（首个标题之前的所有内容）。
    围栏代码块内类似`#`的标题符号将被忽略。
    """
    sections: List[Tuple[int, str, str]] = []
    lines = md_text.splitlines()

    cur_level, cur_heading, cur_body = 0, "", []
    in_code = False

    for line in lines:
        if CODE_FENCE_RE.match(line):
            in_code = not in_code
            cur_body.append(line)
            continue

        if not in_code:
            m = HEADING_RE.match(line)
            if m:
                body_text = "\n".join(cur_body).strip("\n")
                if cur_level != 0 or body_text:
                    sections.append((cur_level, cur_heading, body_text))
                cur_level = len(m.group(1))
                cur_heading = m.group(2).strip()
                cur_body = []
                continue

        cur_body.append(line)

    body_text = "\n".join(cur_body).strip("\n")
    if cur_level != 0 or body_text:
        sections.append((cur_level, cur_heading, body_text))

    return sections


def build_breadcrumbs(sections: List[Tuple[int, str, str]]
                      ) -> List[Tuple[List[str], int, str, str]]:
    """为每个部分附加面包屑导航（完整标题路径）"""
    out = []
    stack: List[Tuple[int, str]] = []

    for level, heading, body in sections:
        if level == 0:
            out.append(([], 0, "", body))
            continue
        while stack and stack[-1][0] >= level:
            stack.pop()
        stack.append((level, heading))
        breadcrumb = [h for _, h in stack]
        out.append((breadcrumb, level, heading, body))

    return out


def detect_primary_level(walked) -> int:
    """
       自动检测应该将哪一级标题作为主要章节的切分边界。
       启发式规则（Heuristic）：选取出现次数至少两次的最浅层级（即 # 数量最少、级别最高的标题）。
       如果没有任何一个层级的标题重复出现，则退而求其次，选取当前存在的最浅层级；
       如果文档里压根没有任何标题，则默认降级为 2 级（其实这时候选几级都无所谓了——整篇文档会被当作一个前言大块来处理，后续的下层切分逻辑依然能正常工作）。
    """
    levels = [lvl for _, lvl, _, _ in walked if lvl > 0]
    if not levels:
        return 2
    counter = Counter(levels)
    for lvl in sorted(counter.keys()):
        if counter[lvl] >= 2:
            return lvl
    return min(counter.keys())


def _extract_document_title(sections: List[Tuple[int, str, str]]) -> Optional[str]:
    """提取文档标题：取第一个 H1 标题文本作为全局文档标题。

    对于 MinerU 输出（所有标题都是 H1），这能提供一个全局上下文，
    注入到每个 chunk 的面包屑中，让 Agent 知道这个 chunk 属于哪个产品/文档。
    """
    for level, heading, _ in sections:
        if level == 1 and heading.strip():
            return heading.strip()
    return None


def _make_table_caption(content: str, section_title: str,
                        is_split: bool, idx: int, total: int) -> Optional[str]:
    """为含表格的 chunk 生成表格标题。"""
    if not _contains_table(content):
        return None
    base = section_title or ''
    if is_split and total > 1:
        return f'{base}（子表 {idx}/{total}）'
    return base if base else None


def _merge_tiny_adjacent_chunks(chunks: List[Chunk], max_chars: int,
                                min_chars: int) -> List[Chunk]:
    """贪婪合并相邻的短纯文本 chunk（无表格、无拆分标记）。

    多轮扫描，直到没有可合并的相邻块。
    """
    if len(chunks) <= 1:
        return chunks

    def _merge_one_pass(chs: List[Chunk]) -> tuple[List[Chunk], bool]:
        """一轮扫描，返回 (合并后的列表, 是否发生了合并)。"""
        out: List[Chunk] = []
        i = 0
        changed = False
        merge_suffix = 0

        while i < len(chs):
            cur = chs[i]
            if (not cur.has_table
                    and not cur.is_split_part
                    and cur.char_count < min_chars
                    and i + 1 < len(chs)):
                nxt = chs[i + 1]
                combined = cur.char_count + len('\n\n') + nxt.char_count
                if (not nxt.has_table
                        and not nxt.is_split_part
                        and combined <= max_chars):
                    merge_suffix += 1
                    merged_content = cur.content + '\n\n' + nxt.content
                    merged_heading = (f'{cur.section_title}；{nxt.section_title}'
                                      if nxt.section_title else cur.section_title)
                    bc = cur.breadcrumb
                    ctx = " > ".join(bc) if bc else "(preamble)"
                    out.append(Chunk(
                        chunk_id=f'{cur.chunk_id}_m{merge_suffix}',
                        source=cur.source, breadcrumb=bc,
                        section_title=merged_heading,
                        heading_level=cur.heading_level,
                        content=merged_content,
                        content_with_context=f'# {ctx}\n\n{merged_content}',
                        char_count=combined,
                        est_token_count=count_tokens(merged_content),
                        has_table=False,
                    ))
                    i += 2
                    changed = True
                    continue
            out.append(cur)
            i += 1
        return out, changed

    # 循环合并直到无法再合并
    changed = True
    result = chunks
    while changed:
        result, changed = _merge_one_pass(result)
        if not changed:
            break

    # 重新编号
    for idx, c in enumerate(result, start=1):
        c.chunk_id = re.sub(r'_m\d+$', '', c.chunk_id)
        c.chunk_id = re.sub(r'_\d{4,}$', f'_{idx:04d}', c.chunk_id)

    return result


# ============================================================================
# 5. 表格检测与按行切分
# ============================================================================

def _is_table_separator(line: str) -> bool:
    s = line.strip().strip('|').strip()
    if not s:
        return False
    parts = [p.strip() for p in s.split('|')]
    return len(parts) >= 2 and all(re.fullmatch(r':?-{2,}:?', p) for p in parts if p)


def split_text_into_blocks(text: str) -> List[Tuple[str, str]]:
    """将章节正文拆分为[(类型, 内容), …]，其中类型为“表格”或“文本”。"""
    lines = text.split('\n')
    blocks: List[Tuple[str, str]] = []
    i, n = 0, len(lines)

    while i < n:
        if '|' in lines[i] and i + 1 < n and _is_table_separator(lines[i + 1]):
            j = i + 2
            while j < n and '|' in lines[j] and lines[j].strip():
                j += 1
            blocks.append(('table', '\n'.join(lines[i:j])))
            i = j
        else:
            j = i
            while j < n:
                if (j + 1 < n and '|' in lines[j]
                        and _is_table_separator(lines[j + 1])):
                    break
                j += 1
            prose = '\n'.join(lines[i:j]).strip('\n')
            if prose:
                blocks.append(('prose', prose))
            i = j

    return blocks


def split_table_by_rows(table_text: str, max_chars: int) -> List[str]:
    """按数据行拆分单个表格；将表头及分隔符复制到每个拆分后的表格中。"""
    lines = [ln for ln in table_text.split('\n') if ln.strip()]
    if len(lines) < 3:
        return [table_text]

    header, sep = lines[0], lines[1]
    data_rows = lines[2:]
    head_size = len(header) + len(sep) + 2

    out: List[str] = []
    cur, cur_size = [header, sep], head_size
    for row in data_rows:
        row_size = len(row) + 1
        if cur_size + row_size > max_chars and len(cur) > 2:
            out.append('\n'.join(cur))
            cur, cur_size = [header, sep], head_size
        cur.append(row)
        cur_size += row_size

    if len(cur) > 2:
        out.append('\n'.join(cur))

    return out


# ============================================================================
# 6. 递归字符拆分（文本形式）
# ============================================================================

def recursive_char_split(text: str, max_chars: int, overlap: int,
                         seps: List[str] = RECURSIVE_SEPARATORS) -> List[str]:
    """尝试按由粗到细的顺序使用分隔符，直至片段符合最大字符数要求，随后带重叠部分进行合并。"""
    if len(text) <= max_chars:
        return [text]

    for sep in seps:
        if sep == "":
            return _hard_char_split(text, max_chars, overlap)
        if sep not in text:
            continue
        pieces = _split_with_sep(text, sep)
        refined: List[str] = []
        for p in pieces:
            if len(p) > max_chars:
                refined.extend(recursive_char_split(
                    p, max_chars, overlap, seps[seps.index(sep) + 1:]))
            else:
                refined.append(p)
        return _merge_with_overlap(refined, max_chars, overlap)

    return _hard_char_split(text, max_chars, overlap)


def _split_with_sep(text: str, sep: str) -> List[str]:
    """使用分隔符`sep`分割文本，并将`sep`保留在每个片段的末尾。

       适用于任意长度的分隔符（单个字符、多个字符、换行符）。
       空片段将被舍弃。
    """
    if not sep:
        return [text]
    parts = text.split(sep)
    return [p + (sep if i < len(parts) - 1 else "")
            for i, p in enumerate(parts) if p != ""]


def _merge_with_overlap(pieces: List[str], max_chars: int, overlap: int) -> List[str]:
    """以贪心方式合并小片段至最大字符数，块之间保留重叠部分。"""
    out: List[str] = []
    cur = ""
    for p in pieces:
        if not cur:
            cur = p
        elif len(cur) + len(p) <= max_chars:
            cur += p
        else:
            out.append(cur)
            tail = cur[-overlap:] if overlap and len(cur) > overlap else ""
            cur = tail + p
    if cur:
        out.append(cur)
    return out


def _hard_char_split(text: str, max_chars: int, overlap: int) -> List[str]:
    out, i = [], 0
    while i < len(text):
        out.append(text[i:i + max_chars])
        i += max(1, max_chars - overlap)
    return out


# ============================================================================
# 7. 把单个 section 切成 N 个 chunk content
# ============================================================================

def _looks_like_table_caption(text: str) -> bool:
    """检测一段简短散文文本是否为后续表格的说明文字。

    规则（满足任意一条即可）：
    a) 以英文冒号:或中文冒号：结尾——明确的“引出下文”标识；
    b) 行数≤2行、字符数≤100个，且不以句末标点结尾。

    相较于原始规则略有放宽（字符数上限由80个调整为100个；分号；纳入句末标点范畴），以避免英文表格说明文字被错误地反向归入前一个表格。
    """
    s = text.strip()
    if not s:
        return False
    if s.rstrip().endswith((':', '：')):
        return True
    lines = [ln for ln in s.split('\n') if ln.strip()]
    if len(lines) > 2:
        return False
    if len(s) > 100:
        return False
    last_char = s.rstrip()[-1]
    if last_char in '。！？；.!?;':
        return False
    return True


def chunk_section_body(body: str, max_chars: int, overlap: int) -> List[Tuple[str, bool]]:
    """将章节正文切分为[(内容, 是否含表格), …]片段。

    1) 拆分为正文文本块/表格块
    2) 对仍过大的内容进一步拆分
    3) 贪心合并相邻的小块，同时遵循一项关键保障规则：
    若当前正文文本块看似是**下一个**表格的标题说明，
    则先清空缓冲区，确保标题与对应表格绑定在一起。
    """
    if not body.strip():
        return []
    if len(body) <= max_chars:
        return [(body, _contains_table(body))]

    expanded: List[Tuple[str, str]] = []
    for kind, blk in split_text_into_blocks(body):
        if len(blk) <= max_chars:
            expanded.append((kind, blk))
        elif kind == 'table':
            for sub in split_table_by_rows(blk, max_chars):
                expanded.append(('table', sub))
        else:
            for sub in recursive_char_split(blk, max_chars, overlap):
                expanded.append(('prose', sub))

    out: List[Tuple[str, bool]] = []
    cur_parts: List[str] = []
    cur_size = 0
    cur_has_table = False
    SEP = "\n\n"
    n_expanded = len(expanded)

    for idx, (kind, blk) in enumerate(expanded):
        add_size = len(blk) + (len(SEP) if cur_parts else 0)

        is_table_caption = (
            kind == 'prose'
            and idx + 1 < n_expanded
            and expanded[idx + 1][0] == 'table'
            and _looks_like_table_caption(blk)
        )
        need_flush = cur_parts and (
            (cur_size + add_size > max_chars) or is_table_caption
        )
        if need_flush:
            out.append((SEP.join(cur_parts), cur_has_table))
            cur_parts = []
            cur_size = 0
            cur_has_table = False
            add_size = len(blk)

        cur_parts.append(blk)
        cur_size += add_size
        if kind == 'table':
            cur_has_table = True

    if cur_parts:
        out.append((SEP.join(cur_parts), cur_has_table))

    return out


def _contains_table(text: str) -> bool:
    lines = text.split('\n')
    for i in range(len(lines) - 1):
        if '|' in lines[i] and _is_table_separator(lines[i + 1]):
            return True
    return False


# ============================================================================
# 8. 顶层入口: 文本 -> [Chunk]
# ============================================================================

def chunk_markdown(md_text: str,
                   source: str,
                   max_chars: int = DEFAULT_MAX_CHARS,
                   overlap: int = DEFAULT_OVERLAP_CHARS,
                   id_prefix: str = "chunk",
                   primary_level: Optional[int] = None,
                   min_chars: int = DEFAULT_MIN_CHARS) -> List[Chunk]:
    """端到端流程。主层级 = None → 自动检测"""
    md_text = normalize_markdown_tables(md_text)
    sections = parse_sections(md_text)
    walked = build_breadcrumbs(sections)

    if primary_level is None:
        primary_level = detect_primary_level(walked)

    # 提取文档标题，注入到所有 chunk 的面包屑中
    doc_title = _extract_document_title(sections)

    merged = _merge_small_primary_sections(walked, max_chars, primary_level)

    chunks: List[Chunk] = []
    counter = 0
    for breadcrumb, level, heading, body in merged:
        if not body.strip():
            continue
        # 将文档标题注入面包屑头部（如果尚未包含）
        if doc_title and (not breadcrumb or breadcrumb[0] != doc_title):
            breadcrumb = [doc_title] + breadcrumb

        parts = chunk_section_body(body, max_chars, overlap)
        total = len(parts)
        for idx, (content, has_table) in enumerate(parts, start=1):
            counter += 1
            ctx_header = " > ".join(breadcrumb) if breadcrumb else "(preamble)"
            content_with_context = f"# {ctx_header}\n\n{content}"
            table_caption = _make_table_caption(
                content, heading, (total > 1), idx, total
            )
            chunks.append(Chunk(
                chunk_id=f"{id_prefix}_{counter:04d}",
                source=source,
                breadcrumb=breadcrumb,
                section_title=heading,
                heading_level=level,
                content=content,
                content_with_context=content_with_context,
                char_count=len(content),
                est_token_count=count_tokens(content_with_context),
                has_table=has_table,
                table_caption=table_caption,
                is_split_part=(total > 1),
                split_part_index=idx if total > 1 else None,
                split_part_total=total if total > 1 else None,
            ))

    # 合并相邻的短纯文本 chunk
    chunks = _merge_tiny_adjacent_chunks(chunks, max_chars, min_chars)
    return chunks


def _merge_small_primary_sections(walked, max_chars: int, primary_level: int):
    """如果一个主章节的全部正文（包含所有深层子章节）字符数不超过最大字符数，则将其合并为单个分块单元。
    否则，将其拆分为独立单元（主章节自身正文 + 各个深层子章节）。

    主层级为可配置参数（原版中该值固定为2）。
    """
    if not walked:
        return []

    primary_indices = [i for i, (_, lvl, _, _) in enumerate(walked)
                       if lvl == primary_level]

    out = []
    # 保留第一个主要章节之前的所有内容（序言+所有层级较浅的内容
    # 位于主层级之上的标题（例如带有介绍性文本的一级标题H1）。
    first_primary = primary_indices[0] if primary_indices else len(walked)
    for item in walked[:first_primary]:
        if item[1] == 0 and not item[3].strip():
            continue
        out.append(item)

    for k, start in enumerate(primary_indices):
        end = primary_indices[k + 1] if k + 1 < len(primary_indices) else len(walked)
        group = walked[start:end]

        p_breadcrumb, p_level, p_heading, p_body = group[0]
        merged_body_parts = []
        if p_body.strip():
            merged_body_parts.append(p_body)
        for sub in group[1:]:
            _, sub_lvl, sub_h, sub_b = sub
            hashes = "#" * sub_lvl
            merged_body_parts.append(f"{hashes} {sub_h}\n{sub_b}".rstrip())
        merged_body = "\n\n".join(merged_body_parts).strip()

        if len(merged_body) <= max_chars or len(group) == 1:
            out.append((p_breadcrumb, p_level, p_heading, merged_body))
        else:
            if p_body.strip():
                out.append((p_breadcrumb, p_level, p_heading, p_body))
            for sub in group[1:]:
                out.append(sub)

    return out


# ============================================================================
# 9. Output writers
# ============================================================================

def write_jsonl(chunks: List[Chunk], path: str) -> None:
    with open(path, "w", encoding="utf-8") as f:
        for c in chunks:
            f.write(json.dumps(asdict(c), ensure_ascii=False) + "\n")


def write_xlsx(chunks: List[Chunk], path: str) -> None:
    if not HAS_OPENPYXL:
        raise ImportError("xlsx output requires openpyxl: pip install openpyxl")

    wb = Workbook()
    ws = wb.active
    ws.title = "chunks"

    headers = [
        "chunk_id", "source", "breadcrumb", "section_title", "heading_level",
        "char_count", "est_token_count", "has_table", "table_caption",
        "is_split_part", "split_part_index", "split_part_total",
        "content", "content_with_context",
    ]
    ws.append(headers)

    for c in chunks:
        ws.append([
            c.chunk_id, c.source, " > ".join(c.breadcrumb), c.section_title,
            c.heading_level, c.char_count, c.est_token_count, c.has_table,
            c.table_caption or "",
            c.is_split_part,
            c.split_part_index if c.split_part_index is not None else "",
            c.split_part_total if c.split_part_total is not None else "",
            c.content, c.content_with_context,
        ])

    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", start_color="305496")
    header_align = Alignment(horizontal="center", vertical="center")
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    widths = {
        "chunk_id": 16, "source": 30, "breadcrumb": 50, "section_title": 26,
        "heading_level": 8, "char_count": 9, "est_token_count": 10,
        "has_table": 9, "is_split_part": 10,
        "split_part_index": 9, "split_part_total": 9,
        "content": 60, "content_with_context": 70,
    }
    for i, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(h, 15)

    wrap_cols = {headers.index(h) + 1 for h in
                 ("breadcrumb", "content", "content_with_context")}
    body_font = Font(name="Arial", size=10)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = body_font
            if cell.column in wrap_cols:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            else:
                cell.alignment = Alignment(vertical="top")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(path)


def _md_anchor(chunk_id: str) -> str:
    return chunk_id.lower().replace("_", "-")


def write_md(chunks: List[Chunk], path: str, min_chars: int = DEFAULT_MIN_CHARS) -> None:
    if not chunks:
        with open(path, "w", encoding="utf-8") as f:
            f.write("# Chunks: (empty)\n\nNo chunks produced.\n")
        return

    n = len(chunks)
    cc = [c.char_count for c in chunks]
    tc = [c.est_token_count for c in chunks]
    n_table = sum(1 for c in chunks if c.has_table)
    n_split = sum(1 for c in chunks if c.is_split_part)
    n_tiny = sum(1 for c in chunks if c.char_count < min_chars)
    source = chunks[0].source

    lines: List[str] = []
    lines.append(f"# 切分输出: {source}")
    lines.append("")
    lines.append("> 本文档供切分质量审阅. 每个 `##` 标题是一个 chunk, 内容按原 markdown 渲染.")
    lines.append("> 若发现表格被切残 / 引导句被吸到上一块 / 块之间语义断裂, 请反馈以调整切分参数.")
    lines.append("")

    lines.append("## 总览统计")
    lines.append("")
    lines.append(f"- **总块数**: {n}")
    lines.append(f"- **字符数** (min / avg / max): {min(cc)} / {sum(cc) // n} / {max(cc)}")
    lines.append(f"- **估算 token** (min / avg / max): {min(tc)} / {sum(tc) // n} / {max(tc)}")
    lines.append(f"- **含表格的块数**: {n_table} ({n_table * 100 // n}%)")
    lines.append(f"- **被进一步切开的子块数**: {n_split}")
    lines.append(f"- **过短 (<{min_chars} 字符) 的块数**: {n_tiny}  ← 仅告警, 不强制合并")
    lines.append("")

    lines.append("## 目录")
    lines.append("")
    for i, c in enumerate(chunks, start=1):
        bc = " > ".join(c.breadcrumb) if c.breadcrumb else "(前言)"
        flag = " 📊" if c.has_table else ""
        split_note = ""
        if c.is_split_part:
            split_note = f" *(子块 {c.split_part_index}/{c.split_part_total})*"
        anchor = _md_anchor(c.chunk_id)
        lines.append(f"{i}. [`{c.chunk_id}`](#{anchor}){flag} — {bc}{split_note}")
    lines.append("")

    for i, c in enumerate(chunks, start=1):
        bc = " > ".join(c.breadcrumb) if c.breadcrumb else "(前言)"
        flag = " 📊" if c.has_table else ""
        anchor = _md_anchor(c.chunk_id)

        lines.append("---")
        lines.append("")
        lines.append(f'<a id="{anchor}"></a>')
        lines.append(f"## 块 {i} / {n}: `{c.chunk_id}`{flag}")
        lines.append("")

        meta_bits = [
            f"**路径**: {bc}",
            f"**章节**: {c.section_title or '(无)'} (H{c.heading_level})",
            f"**{c.char_count} 字符 ≈ {c.est_token_count} tokens**",
        ]
        if c.is_split_part:
            meta_bits.append(f"**子块**: {c.split_part_index} / {c.split_part_total}")
        if c.has_table:
            meta_bits.append("**含表格**")
        if c.table_caption:
            meta_bits.append(f"**表格**: {c.table_caption}")
        lines.append(" · ".join(meta_bits))
        lines.append("")
        lines.append(c.content.rstrip())
        lines.append("")

    with open(path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))


def print_summary(chunks: List[Chunk], min_chars: int = DEFAULT_MIN_CHARS) -> None:
    n = len(chunks)
    if n == 0:
        print("  [WARN] 0 chunks produced.")
        return
    cc = [c.char_count for c in chunks]
    tc = [c.est_token_count for c in chunks]
    n_table = sum(1 for c in chunks if c.has_table)
    n_split = sum(1 for c in chunks if c.is_split_part)
    n_tiny = sum(1 for c in chunks if c.char_count < min_chars)
    print(f"  Total chunks       : {n}")
    print(f"  Chars  min/avg/max : {min(cc)} / {sum(cc) // n} / {max(cc)}")
    print(f"  Tokens min/avg/max : {min(tc)} / {sum(tc) // n} / {max(tc)}")
    print(f"  With table         : {n_table} ({n_table * 100 // n}%)")
    print(f"  Split sub-chunks   : {n_split}")
    print(f"  Tiny (<{min_chars} chars): {n_tiny}  (warning only, not merged)")


# ============================================================================
# 10. 命令行界面/批处理驱动程序
# ============================================================================

def _sanitize_for_id(name: str) -> str:
    """文件基名 -> 分块标识符前缀。保留中日韩文字符，将其他标点符号替换为下划线。"""
    s = re.sub(r'[^A-Za-z0-9\u4e00-\u9fff]+', '_', name)
    return s.strip('_').lower() or "chunk"


def _expand_inputs(patterns: List[str]) -> List[str]:
    """扩展通配符并进行验证，去重同时保留顺序。"""
    out: List[str] = []
    seen = set()
    for p in patterns:
        matches = glob.glob(p, recursive=True) if any(c in p for c in '*?[') else [p]
        if not matches:
            print(f"[WARN] no match for: {p}", file=sys.stderr)
            continue
        for m in matches:
            ap = os.path.abspath(m)
            if ap in seen:
                continue
            if not os.path.isfile(m):
                print(f"[WARN] not a file: {m}", file=sys.stderr)
                continue
            seen.add(ap)
            out.append(m)
    return out


def process_file(input_path: str,
                 out_dir: Optional[str],
                 max_chars: int,
                 overlap: int,
                 min_chars: int,
                 source: Optional[str],
                 id_prefix: Optional[str],
                 primary_level: Optional[int],
                 formats: List[str],
                 pdf_path: Optional[str] = None) -> None:
    basename = os.path.splitext(os.path.basename(input_path))[0]
    derived = _sanitize_for_id(basename)
    src = source or basename
    prefix = id_prefix or derived

    with open(input_path, "r", encoding="utf-8") as f:
        md_text = f.read()

    chunks = chunk_markdown(md_text,
                            source=src,
                            max_chars=max_chars,
                            overlap=overlap,
                            id_prefix=prefix,
                            primary_level=primary_level,
                            min_chars=min_chars)

    target_dir = out_dir or os.path.dirname(os.path.abspath(input_path))
    os.makedirs(target_dir, exist_ok=True)
    base_out = os.path.join(target_dir, basename)

    print(f"\n=== {input_path} ===")
    detected = primary_level if primary_level is not None else \
        detect_primary_level(build_breadcrumbs(parse_sections(normalize_markdown_tables(md_text))))
    print(f"  source={src!r}  id_prefix={prefix!r}  primary_level=H{detected}")

    jsonl_path = None
    if "jsonl" in formats:
        if pdf_path:
            # 有 PDF 时直接出 paged JSONL，跳过中间产物
            paged_out = base_out + "_paged.jsonl"
            # 先写临时 JSONL，回标页码后删掉
            tmp_path = base_out + ".chunks.jsonl"
            write_jsonl(chunks, tmp_path)
            try:
                from add_pages import add_pages as add_pages_to_jsonl
                add_pages_to_jsonl(tmp_path, pdf_path, paged_out)
                os.remove(tmp_path)
                print(f"  [OK]  PAGED -> {paged_out}")
            except Exception as e:
                # 回标失败时保留中间产物
                print(f"  [WARN] 页码回标失败: {e}，保留 -> {tmp_path}")
        else:
            jsonl_path = base_out + ".chunks.jsonl"
            write_jsonl(chunks, jsonl_path)
            print(f"  [OK]  JSONL -> {jsonl_path}")
    if "xlsx" in formats:
        if HAS_OPENPYXL:
            path = base_out + ".chunks.xlsx"
            write_xlsx(chunks, path)
            print(f"  [OK]  XLSX  -> {path}")
        else:
            print("  [SKIP] xlsx (openpyxl not installed; pip install openpyxl)")
    if "md" in formats:
        path = base_out + ".chunks.md"
        write_md(chunks, path, min_chars=min_chars)
        print(f"  [OK]  MD    -> {path}")

    print_summary(chunks, min_chars=min_chars)


def main():
    parser = argparse.ArgumentParser(
        description="具备表格识别能力的通用Markdown分块器"
    )
    parser.add_argument(
        "inputs", nargs="+",
        help="One or more .md files or glob patterns (e.g. 'docs/*.md', 'src/**/*.md').")
    parser.add_argument("--out-dir", default=None,
                        help="Output directory (default: alongside each input file).")
    parser.add_argument("--max-chars", type=int, default=DEFAULT_MAX_CHARS,
                        help=f"Max chars per chunk (default: {DEFAULT_MAX_CHARS}).")
    parser.add_argument("--min-chars", type=int, default=DEFAULT_MIN_CHARS,
                        help=f"Warning threshold for short chunks (default: {DEFAULT_MIN_CHARS}).")
    parser.add_argument("--overlap", type=int, default=DEFAULT_OVERLAP_CHARS,
                        help=f"Overlap chars between prose sub-chunks (default: {DEFAULT_OVERLAP_CHARS}).")
    parser.add_argument("--primary-level", default="auto",
                        choices=["auto", "1", "2", "3", "4", "5", "6"],
                        help="Heading level used as primary section boundary (default: auto-detect).")
    parser.add_argument("--source", default=None,
                        help="Override 'source' field (default: derived from filename).")
    parser.add_argument("--id-prefix", default=None,
                        help="Override chunk_id prefix (default: derived from filename).")
    parser.add_argument("--formats", nargs="+", default=["jsonl", "xlsx", "md"],
                        choices=["jsonl", "xlsx", "md"],
                        help="Which output formats to write (default: all three).")
    parser.add_argument("--pdf", default=None,
                        help="原始 PDF 路径。指定后自动回标页码，输出 _paged.jsonl。")
    args = parser.parse_args()

    files = _expand_inputs(args.inputs)
    if not files:
        print("[ERROR] no input files matched.", file=sys.stderr)
        sys.exit(1)

    pl = None if args.primary_level == "auto" else int(args.primary_level)

    for fp in files:
        try:
            process_file(
                input_path=fp,
                out_dir=args.out_dir,
                max_chars=args.max_chars,
                overlap=args.overlap,
                min_chars=args.min_chars,
                source=args.source,
                id_prefix=args.id_prefix,
                primary_level=pl,
                formats=args.formats,
                pdf_path=args.pdf,
            )
        except Exception as e:
            print(f"[ERROR] failed on {fp}: {e}", file=sys.stderr)


if __name__ == "__main__":
    main()
